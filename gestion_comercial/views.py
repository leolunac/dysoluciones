from pathlib import Path
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse
from django.utils import timezone

from .forms import (
    LiquidacionForm,
    DetalleLiquidacionForm,
    CatalogoPrecioForm,
    CotizacionForm,
    DatosComercialesCotizacionForm,
    DetalleCotizacionForm,
)
from .models import (
    Liquidacion,
    CatalogoPrecio,
    DetalleLiquidacion,
    Cotizacion,
    DetalleCotizacion,
    ImagenCotizacion,
)


# =========================================================
# PERFILES / SEGURIDAD GESTIÓN COMERCIAL
# =========================================================

GRUPO_AUXILIAR = "GESTION_AUXILIAR"
GRUPO_COORDINADOR = "GESTION_COORDINADOR"
GRUPO_FACTURACION = "GESTION_FACTURACION"
GRUPO_GERENCIA = "GESTION_GERENCIA"
GRUPO_SUPERVISOR = "GESTION_SUPERVISOR"


def _pertenece(user, *grupos):
    return (
        user.is_superuser
        or user.groups.filter(name__in=grupos).exists()
    )


def _exigir(user, *grupos):
    if not _pertenece(user, *grupos):
        raise PermissionDenied


def _puede_elaborar(user):
    return _pertenece(user, GRUPO_AUXILIAR)


def _puede_revisar(user):
    return _pertenece(user, GRUPO_COORDINADOR)


def _puede_facturar(user):
    return _pertenece(user, GRUPO_FACTURACION)


def _puede_consultar(user):
    return _pertenece(
        user,
        GRUPO_AUXILIAR,
        GRUPO_COORDINADOR,
        GRUPO_FACTURACION,
        GRUPO_GERENCIA,
    )


@login_required
def panel_gestion_comercial(request):
    _exigir(
        request.user,
        GRUPO_AUXILIAR,
        GRUPO_COORDINADOR,
        GRUPO_FACTURACION,
        GRUPO_GERENCIA,
    )

    liquidaciones = (
        Liquidacion.objects
        .select_related(
            "cliente",
            "creado_por",
            "revisado_por",
        )
        .all()
    )

    contexto = {
        "liquidaciones": liquidaciones,
        "total": liquidaciones.count(),
        "revision": liquidaciones.filter(estado="REVISION").count(),
        "devueltas": liquidaciones.filter(estado="DEVUELTA").count(),
        "aprobadas": liquidaciones.filter(estado="APROBADA").count(),
        "listas_facturar": liquidaciones.filter(
            estado="LISTA_FACTURAR"
        ).count(),
        "facturadas": liquidaciones.filter(estado="FACTURADA").count(),
        "puede_elaborar": _puede_elaborar(request.user),
        "puede_revisar": _puede_revisar(request.user),
        "puede_facturar": _puede_facturar(request.user),
        "puede_consultar": _puede_consultar(request.user),
        "puede_ver_consolidado": _pertenece(
            request.user,
            GRUPO_FACTURACION,
            GRUPO_GERENCIA,
        ),
    }

    return render(
        request,
        "gestion_comercial/panel.html",
        contexto,
    )


@login_required
def nueva_liquidacion(request):
    _exigir(request.user, GRUPO_AUXILIAR)

    if request.method == "POST":
        form = LiquidacionForm(request.POST)

        if form.is_valid():
            liquidacion = form.save(commit=False)
            liquidacion.creado_por = request.user
            liquidacion.estado = "BORRADOR"
            liquidacion.save()

            return redirect(
                "gestion_comercial:editar_liquidacion",
                liquidacion_id=liquidacion.id,
            )
    else:
        form = LiquidacionForm()

    return render(
        request,
        "gestion_comercial/nueva_liquidacion.html",
        {"form": form},
    )


@login_required
def editar_liquidacion(request, liquidacion_id):
    _exigir(
        request.user,
        GRUPO_AUXILIAR,
        GRUPO_COORDINADOR,
        GRUPO_FACTURACION,
        GRUPO_GERENCIA,
    )

    liquidacion = get_object_or_404(
        Liquidacion.objects.select_related(
            "cliente",
            "creado_por",
            "revisado_por",
        ),
        id=liquidacion_id,
    )

    if request.method == "POST":
        _exigir(request.user, GRUPO_AUXILIAR)

        if (
            not request.user.is_superuser
            and liquidacion.creado_por_id != request.user.id
        ):
            raise PermissionDenied

        if liquidacion.estado not in ["BORRADOR", "DEVUELTA"]:
            raise PermissionDenied

        form_detalle = DetalleLiquidacionForm(request.POST)

        if form_detalle.is_valid():
            detalle = form_detalle.save(commit=False)
            detalle.liquidacion = liquidacion

            if detalle.tipo == "MATERIAL" and detalle.catalogo:
                detalle.descripcion = detalle.catalogo.descripcion

            if detalle.tipo != "MATERIAL":
                detalle.catalogo = None

            detalle.save()
            liquidacion.recalcular_totales()

            return redirect(
                "gestion_comercial:editar_liquidacion",
                liquidacion_id=liquidacion.id,
            )
    else:
        form_detalle = DetalleLiquidacionForm()

    detalles = (
        liquidacion.detalles
        .select_related("catalogo")
        .all()
        .order_by("id")
    )

    catalogos = (
        CatalogoPrecio.objects
        .filter(activo=True)
        .order_by("descripcion")
    )

    return render(
        request,
        "gestion_comercial/editar_liquidacion.html",
        {
            "liquidacion": liquidacion,
            "form_detalle": form_detalle,
            "detalles": detalles,
            "catalogos": catalogos,
            "form_catalogo": CatalogoPrecioForm(),
            "puede_elaborar": _puede_elaborar(request.user),
            "puede_revisar": _puede_revisar(request.user),
            "puede_facturar": _puede_facturar(request.user),
            "puede_consultar": _puede_consultar(request.user),
        },
    )


@login_required
def crear_accesorio(request, liquidacion_id):
    _exigir(request.user, GRUPO_AUXILIAR)

    liquidacion = get_object_or_404(
        Liquidacion,
        id=liquidacion_id,
    )

    if (
        not request.user.is_superuser
        and liquidacion.creado_por_id != request.user.id
    ):
        raise PermissionDenied

    if liquidacion.estado not in ["BORRADOR", "DEVUELTA"]:
        raise PermissionDenied

    if request.method != "POST":
        return redirect(
            "gestion_comercial:editar_liquidacion",
            liquidacion_id=liquidacion.id,
        )

    form_catalogo = CatalogoPrecioForm(request.POST)

    if form_catalogo.is_valid():
        accesorio = form_catalogo.save(commit=False)
        accesorio.activo = True
        accesorio.save()

        return redirect(
            "gestion_comercial:editar_liquidacion",
            liquidacion_id=liquidacion.id,
        )

    form_detalle = DetalleLiquidacionForm()
    detalles = (
        liquidacion.detalles
        .select_related("catalogo")
        .all()
        .order_by("id")
    )
    catalogos = (
        CatalogoPrecio.objects
        .filter(activo=True)
        .order_by("descripcion")
    )

    return render(
        request,
        "gestion_comercial/editar_liquidacion.html",
        {
            "liquidacion": liquidacion,
            "form_detalle": form_detalle,
            "detalles": detalles,
            "catalogos": catalogos,
            "form_catalogo": form_catalogo,
            "mostrar_form_catalogo": True,
            "puede_elaborar": _puede_elaborar(request.user),
            "puede_revisar": _puede_revisar(request.user),
            "puede_facturar": _puede_facturar(request.user),
            "puede_consultar": _puede_consultar(request.user),
        },
    )


@login_required
def enviar_revision(request, liquidacion_id):
    _exigir(request.user, GRUPO_AUXILIAR)

    liquidacion = get_object_or_404(
        Liquidacion,
        id=liquidacion_id,
    )

    if (
        not request.user.is_superuser
        and liquidacion.creado_por_id != request.user.id
    ):
        raise PermissionDenied

    if liquidacion.estado not in ["BORRADOR", "DEVUELTA"]:
        raise PermissionDenied

    if request.method == "POST" and liquidacion.detalles.exists():
        liquidacion.estado = "REVISION"
        liquidacion.save(
            update_fields=[
                "estado",
                "fecha_actualizacion",
            ]
        )

    return redirect("gestion_comercial:panel")


@login_required
def eliminar_detalle(request, detalle_id):
    _exigir(request.user, GRUPO_AUXILIAR)

    detalle = get_object_or_404(
        DetalleLiquidacion,
        id=detalle_id,
    )
    liquidacion = detalle.liquidacion

    if (
        not request.user.is_superuser
        and liquidacion.creado_por_id != request.user.id
    ):
        raise PermissionDenied

    if liquidacion.estado not in ["BORRADOR", "DEVUELTA"]:
        return redirect(
            "gestion_comercial:editar_liquidacion",
            liquidacion_id=liquidacion.id,
        )

    if request.method == "POST":
        detalle.delete()

    return redirect(
        "gestion_comercial:editar_liquidacion",
        liquidacion_id=liquidacion.id,
    )


@login_required
def editar_detalle(request, detalle_id):
    _exigir(request.user, GRUPO_AUXILIAR)

    detalle = get_object_or_404(
        DetalleLiquidacion,
        id=detalle_id,
    )
    liquidacion = detalle.liquidacion

    if (
        not request.user.is_superuser
        and liquidacion.creado_por_id != request.user.id
    ):
        raise PermissionDenied

    if liquidacion.estado not in ["BORRADOR", "DEVUELTA"]:
        return redirect(
            "gestion_comercial:editar_liquidacion",
            liquidacion_id=liquidacion.id,
        )

    if request.method == "POST":
        form = DetalleLiquidacionForm(
            request.POST,
            instance=detalle,
        )

        if form.is_valid():
            detalle_editado = form.save(commit=False)

            if (
                detalle_editado.tipo == "MATERIAL"
                and detalle_editado.catalogo
            ):
                detalle_editado.descripcion = (
                    detalle_editado.catalogo.descripcion
                )

            if detalle_editado.tipo != "MATERIAL":
                detalle_editado.catalogo = None

            detalle_editado.save()
            liquidacion.recalcular_totales()

            return redirect(
                "gestion_comercial:editar_liquidacion",
                liquidacion_id=liquidacion.id,
            )
    else:
        form = DetalleLiquidacionForm(instance=detalle)

    catalogos = (
        CatalogoPrecio.objects
        .filter(activo=True)
        .order_by("descripcion")
    )

    return render(
        request,
        "gestion_comercial/editar_detalle.html",
        {
            "liquidacion": liquidacion,
            "detalle": detalle,
            "form": form,
            "catalogos": catalogos,
        },
    )


@login_required
def aprobar_liquidacion(request, liquidacion_id):
    _exigir(request.user, GRUPO_COORDINADOR)

    liquidacion = get_object_or_404(
        Liquidacion,
        id=liquidacion_id,
    )

    if (
        not request.user.is_superuser
        and liquidacion.creado_por_id == request.user.id
    ):
        raise PermissionDenied

    if liquidacion.estado != "REVISION":
        return redirect(
            "gestion_comercial:editar_liquidacion",
            liquidacion_id=liquidacion.id,
        )

    if request.method == "POST":
        liquidacion.estado = "APROBADA"
        liquidacion.revisado_por = request.user
        liquidacion.fecha_revision = timezone.now()
        liquidacion.save(
            update_fields=[
                "estado",
                "revisado_por",
                "fecha_revision",
                "fecha_actualizacion",
            ]
        )

    return redirect(
        "gestion_comercial:editar_liquidacion",
        liquidacion_id=liquidacion.id,
    )


@login_required
def devolver_liquidacion(request, liquidacion_id):
    _exigir(request.user, GRUPO_COORDINADOR)

    liquidacion = get_object_or_404(
        Liquidacion,
        id=liquidacion_id,
    )

    if (
        not request.user.is_superuser
        and liquidacion.creado_por_id == request.user.id
    ):
        raise PermissionDenied

    if liquidacion.estado != "REVISION":
        return redirect(
            "gestion_comercial:editar_liquidacion",
            liquidacion_id=liquidacion.id,
        )

    if request.method == "POST":
        observacion = request.POST.get(
            "observaciones_revision",
            ""
        ).strip()

        if not observacion:
            return redirect(
                "gestion_comercial:editar_liquidacion",
                liquidacion_id=liquidacion.id,
            )

        liquidacion.estado = "DEVUELTA"
        liquidacion.revisado_por = request.user
        liquidacion.fecha_revision = timezone.now()
        liquidacion.observaciones_revision = observacion
        liquidacion.save(
            update_fields=[
                "estado",
                "revisado_por",
                "fecha_revision",
                "observaciones_revision",
                "fecha_actualizacion",
            ]
        )

    return redirect(
        "gestion_comercial:editar_liquidacion",
        liquidacion_id=liquidacion.id,
    )


@login_required
def enviar_facturacion(request, liquidacion_id):
    _exigir(request.user, GRUPO_COORDINADOR)

    liquidacion = get_object_or_404(
        Liquidacion,
        id=liquidacion_id,
    )

    if liquidacion.estado != "APROBADA":
        return redirect(
            "gestion_comercial:editar_liquidacion",
            liquidacion_id=liquidacion.id,
        )

    if request.method == "POST":
        liquidacion.estado = "LISTA_FACTURAR"
        liquidacion.save(
            update_fields=[
                "estado",
                "fecha_actualizacion",
            ]
        )

    return redirect(
        "gestion_comercial:editar_liquidacion",
        liquidacion_id=liquidacion.id,
    )


@login_required
def marcar_facturada(request, liquidacion_id):
    _exigir(request.user, GRUPO_FACTURACION)

    liquidacion = get_object_or_404(
        Liquidacion,
        id=liquidacion_id,
    )

    if liquidacion.estado not in [
        "APROBADA",
        "LISTA_FACTURAR",
    ]:
        return redirect(
            "gestion_comercial:editar_liquidacion",
            liquidacion_id=liquidacion.id,
        )

    if request.method == "POST":
        numero_factura = request.POST.get(
            "numero_factura",
            ""
        ).strip()

        if not numero_factura:
            return redirect(
                "gestion_comercial:editar_liquidacion",
                liquidacion_id=liquidacion.id,
            )

        liquidacion.numero_factura = numero_factura
        liquidacion.fecha_facturacion = timezone.now()
        liquidacion.estado = "FACTURADA"
        liquidacion.save(
            update_fields=[
                "numero_factura",
                "fecha_facturacion",
                "estado",
                "fecha_actualizacion",
            ]
        )

    return redirect(
        "gestion_comercial:editar_liquidacion",
        liquidacion_id=liquidacion.id,
    )


@login_required
def consolidado_facturacion(request):
    _exigir(
        request.user,
        GRUPO_COORDINADOR,
        GRUPO_FACTURACION,
        GRUPO_GERENCIA,
    )

    from django.db.models import Sum
    from operacion.models import Cliente

    facturadas = (
        Liquidacion.objects
        .filter(estado="FACTURADA")
        .select_related(
            "cliente",
            "creado_por",
        )
    )

    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()
    cliente_id = request.GET.get("cliente", "").strip()

    if fecha_desde:
        facturadas = facturadas.filter(
            fecha_facturacion__date__gte=fecha_desde
        )
    if fecha_hasta:
        facturadas = facturadas.filter(
            fecha_facturacion__date__lte=fecha_hasta
        )
    if cliente_id:
        facturadas = facturadas.filter(
            cliente_id=cliente_id
        )

    facturadas = facturadas.order_by("-fecha_facturacion")

    totales = facturadas.aggregate(
        total_sin_iva=Sum("valor_sin_iva"),
        total_iva=Sum("valor_iva"),
        total_facturado=Sum("valor_total"),
    )

    clientes = (
        Cliente.objects
        .filter(activo=True)
        .order_by("nombre")
    )

    contexto = {
        "facturadas": facturadas,
        "cantidad_facturas": facturadas.count(),
        "total_sin_iva": totales["total_sin_iva"] or 0,
        "total_iva": totales["total_iva"] or 0,
        "total_facturado": totales["total_facturado"] or 0,
        "clientes": clientes,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "cliente_seleccionado": cliente_id,
    }

    return render(
        request,
        "gestion_comercial/consolidado_facturacion.html",
        contexto,
    )


@login_required
def exportar_facturacion_excel(request):
    _exigir(
        request.user,
        GRUPO_COORDINADOR,
        GRUPO_FACTURACION,
        GRUPO_GERENCIA,
    )

    from io import BytesIO
    from django.http import HttpResponse
    from django.db.models import Sum
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    facturadas = (
        Liquidacion.objects
        .filter(estado="FACTURADA")
        .select_related("cliente")
    )

    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()
    cliente_id = request.GET.get("cliente", "").strip()

    if fecha_desde:
        facturadas = facturadas.filter(
            fecha_facturacion__date__gte=fecha_desde
        )
    if fecha_hasta:
        facturadas = facturadas.filter(
            fecha_facturacion__date__lte=fecha_hasta
        )
    if cliente_id:
        facturadas = facturadas.filter(
            cliente_id=cliente_id
        )

    facturadas = facturadas.order_by("-fecha_facturacion")

    totales = facturadas.aggregate(
        total_sin_iva=Sum("valor_sin_iva"),
        total_iva=Sum("valor_iva"),
        total_facturado=Sum("valor_total"),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturación"

    ws.merge_cells("A1:G1")
    ws["A1"] = "SIGOB - Consolidado de Facturación"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    filtros = []
    if fecha_desde:
        filtros.append(f"Desde: {fecha_desde}")
    if fecha_hasta:
        filtros.append(f"Hasta: {fecha_hasta}")
    if cliente_id:
        primera = facturadas.first()
        if primera:
            filtros.append(f"Unidad: {primera.cliente.nombre}")

    ws.merge_cells("A2:G2")
    ws["A2"] = " | ".join(filtros) if filtros else "Todas las facturas"
    ws["A2"].alignment = Alignment(horizontal="center")

    encabezados = [
        "Factura",
        "Fecha",
        "Unidad",
        "Trabajo / concepto",
        "Valor sin IVA",
        "IVA",
        "Total",
    ]

    for columna, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=4, column=columna, value=encabezado)
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center")

    fila = 5

    for liquidacion in facturadas:
        ws.cell(fila, 1, liquidacion.numero_factura or "")
        ws.cell(
            fila,
            2,
            (
                liquidacion.fecha_facturacion.replace(tzinfo=None)
                if liquidacion.fecha_facturacion
                else None
            ),
        )
        ws.cell(fila, 3, liquidacion.cliente.nombre)
        ws.cell(fila, 4, liquidacion.descripcion or "")
        ws.cell(fila, 5, float(liquidacion.valor_sin_iva or 0))
        ws.cell(fila, 6, float(liquidacion.valor_iva or 0))
        ws.cell(fila, 7, float(liquidacion.valor_total or 0))
        fila += 1

    ws.cell(fila, 4, "TOTALES")
    ws.cell(fila, 4).font = Font(bold=True)
    ws.cell(fila, 5, float(totales["total_sin_iva"] or 0))
    ws.cell(fila, 6, float(totales["total_iva"] or 0))
    ws.cell(fila, 7, float(totales["total_facturado"] or 0))

    for col in range(5, 8):
        for row in range(5, fila + 1):
            ws.cell(row, col).number_format = '$#,##0'

    for row in range(5, fila):
        ws.cell(row, 2).number_format = "dd/mm/yyyy hh:mm"

    anchos = {
        1: 18,
        2: 20,
        3: 35,
        4: 45,
        5: 18,
        6: 18,
        7: 18,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[get_column_letter(columna)].width = ancho

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:G{max(fila - 1, 4)}"

    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)

    response = HttpResponse(
        archivo.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        'attachment; filename="consolidado_facturacion.xlsx"'
    )

    return response


# =========================================================
# COTIZACIONES
# =========================================================

@login_required
def lista_cotizaciones(request):
    _exigir(
        request.user,
        GRUPO_AUXILIAR,
        GRUPO_COORDINADOR,
        GRUPO_FACTURACION,
        GRUPO_GERENCIA,
        GRUPO_SUPERVISOR,
    )

    cotizaciones = (
        Cotizacion.objects
        .select_related(
            "cliente",
            "elaborado_por",
            "servicio",
            "actividad",
            "bitacora",
        )
        .all()
    )

    return render(
    request,
    "gestion_comercial/cotizaciones/lista.html",
    {
        "cotizaciones": cotizaciones,
        "es_supervisor": _pertenece(
            request.user,
            GRUPO_SUPERVISOR,
        ),
    },
)


@login_required
def nueva_cotizacion(request):
    _exigir(
        request.user,
        GRUPO_AUXILIAR,
        GRUPO_COORDINADOR,
        GRUPO_FACTURACION,
        GRUPO_GERENCIA,
        GRUPO_SUPERVISOR,
    )

    if request.method == "POST":
        form = CotizacionForm(request.POST)

        if form.is_valid():
            cotizacion = form.save(commit=False)
            cotizacion.elaborado_por = request.user
            cotizacion.estado = "BORRADOR"
            cotizacion.save()

            return redirect(
                "gestion_comercial:editar_cotizacion",
                cotizacion_id=cotizacion.id,
            )
    else:
        form = CotizacionForm()

    return render(
        request,
        "gestion_comercial/cotizaciones/nueva.html",
        {
            "form": form,
        },
    )


@login_required
def editar_cotizacion(request, cotizacion_id):
    _exigir(
        request.user,
        GRUPO_AUXILIAR,
        GRUPO_COORDINADOR,
        GRUPO_FACTURACION,
        GRUPO_GERENCIA,
        GRUPO_SUPERVISOR,
    )

    cotizacion = get_object_or_404(
        Cotizacion.objects.select_related(
            "cliente",
            "elaborado_por",
        ),
        id=cotizacion_id,
    )

    form_comercial = DatosComercialesCotizacionForm(instance=cotizacion)
    form_detalle = DetalleCotizacionForm()

    if request.method == "POST":
        accion = request.POST.get("accion", "agregar_detalle")

        if accion == "guardar_comercial":
            form_comercial = DatosComercialesCotizacionForm(
                request.POST,
                instance=cotizacion,
            )

            if form_comercial.is_valid():
                form_comercial.save()
                cotizacion.recalcular_totales()

                return redirect(
                    "gestion_comercial:editar_cotizacion",
                    cotizacion_id=cotizacion.id,
                )
        elif accion == "agregar_detalle":
            form_detalle = DetalleCotizacionForm(request.POST)

            if form_detalle.is_valid():
                detalle = form_detalle.save(commit=False)
                detalle.cotizacion = cotizacion

                if detalle.tipo == "MATERIAL" and detalle.catalogo:
                    detalle.descripcion = detalle.catalogo.descripcion

                if detalle.tipo != "MATERIAL":
                    detalle.catalogo = None

                detalle.save()
                cotizacion.recalcular_totales()

                return redirect(
                    "gestion_comercial:editar_cotizacion",
                    cotizacion_id=cotizacion.id,
                )

        elif accion == "agregar_imagen":
            imagen = request.FILES.get("imagen")
            descripcion = request.POST.get("descripcion", "").strip()

            if imagen:
                ImagenCotizacion.objects.create(
                    cotizacion=cotizacion,
                    imagen=imagen,
                    descripcion=descripcion,
                )

                return redirect(
                    "gestion_comercial:editar_cotizacion",
                    cotizacion_id=cotizacion.id,
                )
    detalles = (
        cotizacion.detalles
        .select_related("catalogo")
        .all()
        .order_by("id")
    )

    catalogos = (
        CatalogoPrecio.objects
        .filter(activo=True)
        .order_by("descripcion")
    )

    return render(
        request,
        "gestion_comercial/cotizaciones/editar.html",
        {
            "cotizacion": cotizacion,
            "form_comercial": form_comercial,
            "form_detalle": form_detalle,
            "detalles": detalles,
            "catalogos": catalogos,
            "puede_aprobar_cotizacion": _pertenece(
                request.user,
                GRUPO_COORDINADOR,
                GRUPO_GERENCIA,
            ),
        },
    )

@login_required
def finalizar_elaboracion_cotizacion(request, cotizacion_id):
    _exigir(
        request.user,
        GRUPO_AUXILIAR,
        GRUPO_COORDINADOR,
        GRUPO_FACTURACION,
        GRUPO_GERENCIA,
        GRUPO_SUPERVISOR,
    )

    cotizacion = get_object_or_404(
        Cotizacion,
        id=cotizacion_id,
    )

    if request.method == "POST":
        if cotizacion.estado == "BORRADOR" and cotizacion.detalles.exists():
            cotizacion.estado = "ELABORADA"

            if not cotizacion.fecha_emision:
                cotizacion.fecha_emision = timezone.localdate()

            cotizacion.save(
                update_fields=[
                    "estado",
                    "fecha_emision",
                    "fecha_actualizacion",
                ]
            )

    return redirect(
        "gestion_comercial:editar_cotizacion",
        cotizacion_id=cotizacion.id,
    )



@login_required
def generar_pdf_cotizacion(request, cotizacion_id):
    _exigir(
        request.user,
        GRUPO_AUXILIAR,
        GRUPO_COORDINADOR,
        GRUPO_FACTURACION,
        GRUPO_GERENCIA,
        GRUPO_SUPERVISOR,
    )

    from io import BytesIO

    from django.conf import settings
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        Image,
        PageBreak,
        KeepTogether,
    )

    cotizacion = get_object_or_404(
        Cotizacion.objects.select_related(
            "cliente",
            "elaborado_por",
        ),
        id=cotizacion_id,
    )

    if cotizacion.estado == "BORRADOR":
        return redirect(
            "gestion_comercial:editar_cotizacion",
            cotizacion_id=cotizacion.id,
        )

    buffer = BytesIO()

    nombre = (
        cotizacion.numero_cotizacion
        or f"COT-{cotizacion.id}"
    )

    azul = colors.HexColor("#1F4E78")
    azul_claro = colors.HexColor("#5B9BD5")
    gris = colors.HexColor("#4B5563")
    gris_borde = colors.HexColor("#B8C4CE")

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=13 * mm,
        bottomMargin=15 * mm,
        title=nombre,
        author="D&S SOLUCIONES EN BOMBEO S.A.S.",
    )

    styles = getSampleStyleSheet()

    normal = ParagraphStyle(
        "NormalPropuesta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9.4,
        leading=12.8,
        textColor=colors.HexColor("#1F2937"),
        alignment=TA_JUSTIFY,
    )

    normal_left = ParagraphStyle(
        "NormalLeft",
        parent=normal,
        alignment=0,
    )

    small = ParagraphStyle(
        "Small",
        parent=normal,
        fontSize=8.5,
        leading=11,
    )

    small_right = ParagraphStyle(
        "SmallRight",
        parent=small,
        alignment=TA_RIGHT,
    )

    section = ParagraphStyle(
        "Section",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=10.8,
        leading=13,
        textColor=colors.black,
        spaceBefore=8,
        spaceAfter=5,
    )

    title_right = ParagraphStyle(
        "TitleRight",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=15.5,
        leading=17,
        textColor=azul,
        alignment=TA_RIGHT,
        spaceAfter=2,
    )

    number_right = ParagraphStyle(
        "NumberRight",
        parent=normal,
        fontName="Helvetica-Bold",
        fontSize=10.5,
        textColor=azul,
        alignment=TA_RIGHT,
    )

    footer_style = ParagraphStyle(
        "Footer",
        parent=small,
        alignment=TA_CENTER,
        textColor=gris,
    )

    meses = {
        1: "enero",
        2: "febrero",
        3: "marzo",
        4: "abril",
        5: "mayo",
        6: "junio",
        7: "julio",
        8: "agosto",
        9: "septiembre",
        10: "octubre",
        11: "noviembre",
        12: "diciembre",
    }

    fecha = cotizacion.fecha_emision or cotizacion.fecha_creacion.date()
    fecha_larga = (
        f"{fecha.day} de {meses[fecha.month]} de {fecha.year}"
    )

    def money(value):
        return f"$ {value:,.0f}".replace(",", ".")

    def limpiar_lineas(texto):
        if not texto:
            return []
        return [
            linea.strip(" •-\t")
            for linea in str(texto).splitlines()
            if linea.strip()
        ]

    story = []

    # =====================================================
    # ENCABEZADO CORPORATIVO
    # =====================================================

    logo_path = (
        Path(settings.BASE_DIR)
        / "static"
        / "img"
        / "logo_dys.png"
    )

    logo = ""
    if logo_path.exists():
        logo = Image(
            str(logo_path),
            width=44 * mm,
            height=21 * mm,
            kind="proportional",
        )

    encabezado_derecha = [
        Paragraph("PROPUESTA COMERCIAL", title_right),
        Paragraph(nombre, number_right),
    ]

    header = Table(
        [[logo, encabezado_derecha]],
        colWidths=[84 * mm, 96 * mm],
    )
    header.setStyle(
        TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LINEBELOW", (0, 0), (-1, -1), 2.2, azul),
        ])
    )
    story.append(header)
    story.append(Spacer(1, 7))

    # =====================================================
    # DESTINATARIO
    # =====================================================

    story.append(
        Paragraph(
            f"Envigado, {fecha_larga}",
            normal_left,
        )
    )
    story.append(Spacer(1, 6))

 
    contacto = (
        getattr(cotizacion.cliente, "administrador", "")
        or getattr(cotizacion.cliente, "contacto", "")
        or ""
    )
    telefono = (
        getattr(cotizacion.cliente, "telefono", "")
        or getattr(cotizacion.cliente, "telefono_porteria", "")
        or ""
    )

    if contacto and str(contacto).strip().lower() not in {
        "pendiente",
        "por definir",
        "n/a",
        "na",
        "-",
    }:
        story.append(
            Paragraph(f"<b>Contacto:</b> {contacto}", normal_left)
        )

    story.append(
        Paragraph(
            f"<b>{cotizacion.cliente.nombre}</b>",
            normal_left,
        )
    )

    if telefono:
        story.append(
            Paragraph(f"<b>Tel:</b> {telefono}", normal_left)
        )

    story.append(
        Paragraph(
            f"<b>Asunto:</b> {cotizacion.asunto}",
            normal_left,
        )
    )
    story.append(Spacer(1, 6))

    story.append(Paragraph("Respetados Señores:", normal_left))
    story.append(Spacer(1, 6))

    introduccion = (
        "En atención a su gentil invitación, adjunto remitimos la "
        "propuesta técnica y económica relacionada con el desarrollo "
        "de las actividades descritas en el asunto. Quedamos dispuestos "
        "a aclarar y/o complementar la información suministrada."
    )
    story.append(Paragraph(introduccion, normal))
    story.append(Spacer(1, 6))

    # =====================================================
    # 1. OBJETO
    # =====================================================

    story.append(Paragraph("1. OBJETO DE LA PROPUESTA", section))

    objeto = (
        cotizacion.descripcion.strip()
        if cotizacion.descripcion
        else cotizacion.asunto
    )

    story.append(
        Paragraph(
            str(objeto).replace("\n", "<br/>"),
            normal,
        )
    )
    story.append(Spacer(1, 6))

    # =====================================================
    # 2. ALCANCE
    # =====================================================

    story.append(Paragraph("2. ALCANCE TÉCNICO.", section))

    actividades = limpiar_lineas(cotizacion.alcance_tecnico)

    if actividades:
        for actividad in actividades:
            story.append(
                Paragraph(
                    f"• {actividad}",
                    normal_left,
                )
            )
    else:
        story.append(
            Paragraph(
                "Alcance técnico pendiente por definir.",
                normal_left,
            )
        )

    story.append(Spacer(1, 5))

    # =====================================================
    # 3. OFERTA ECONÓMICA
    # =====================================================

    story.append(Paragraph("3. OFERTA ECONÓMICA.", section))

    concepto = (
        cotizacion.concepto_comercial.strip()
        if cotizacion.concepto_comercial
        else cotizacion.asunto
    )

    oferta = [
        [
            Paragraph("<b>DESCRIPCIÓN</b>", small),
            Paragraph("<b>CANTIDAD</b>", small),
            Paragraph("<b>VALOR<br/>UNITARIO</b>", small),
            Paragraph("<b>VALOR<br/>TOTAL</b>", small),
        ],
        [
            Paragraph(
                str(concepto).replace("\n", "<br/>"),
                normal_left,
            ),
            Paragraph("1", small),
            Paragraph(
                money(cotizacion.valor_sin_iva),
                small_right,
            ),
            Paragraph(
                money(cotizacion.valor_sin_iva),
                small_right,
            ),
        ],
    ]

    tabla_oferta = Table(
        oferta,
        colWidths=[92 * mm, 22 * mm, 33 * mm, 33 * mm],
    )
    tabla_oferta.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), azul_claro),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.75, colors.black),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ])
    )
    story.append(tabla_oferta)

    resumen = [
        [
            "",
            Paragraph("<b>SUBTOTAL</b>", small_right),
            Paragraph(
                f"<b>{money(cotizacion.valor_sin_iva)}</b>",
                small_right,
            ),
        ],
        [
            "",
            Paragraph(
                f"<b>IVA ({cotizacion.porcentaje_iva:.0f}%)</b>",
                small_right,
            ),
            Paragraph(
                f"<b>{money(cotizacion.valor_iva)}</b>",
                small_right,
            ),
        ],
        [
            "",
            Paragraph("<b>TOTAL</b>", small_right),
            Paragraph(
                f"<b>{money(cotizacion.valor_total)}</b>",
                small_right,
            ),
        ],
    ]

    tabla_resumen = Table(
        resumen,
        colWidths=[92 * mm, 55 * mm, 33 * mm],
    )
    tabla_resumen.setStyle(
        TableStyle([
            ("SPAN", (0, 0), (0, -1)),
            ("GRID", (1, 0), (-1, -1), 0.75, colors.black),
            ("BACKGROUND", (1, -1), (-1, -1), azul_claro),
            ("TOPPADDING", (1, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (1, 0), (-1, -1), 4),
        ])
    )
    story.append(tabla_resumen)
    story.append(Spacer(1, 8))

    # =====================================================
    # 4. CONSIDERACIONES
    # =====================================================

    story.append(
        Paragraph(
            "4. CONSIDERACIONES A TENER EN CUENTA EN ESTA OFERTA ECONÓMICA.",
            section,
        )
    )

    consideraciones = limpiar_lineas(cotizacion.observaciones)

    for item in consideraciones:
        story.append(
            Paragraph(
                f"• {item}",
                normal_left,
            )
        )

    story.append(
        Paragraph(
            f"• La propuesta tiene validez por "
            f"{cotizacion.vigencia_dias} días calendario.",
            normal_left,
        )
    )

    if cotizacion.forma_pago:
        story.append(
            Paragraph(
                f"• <b>Forma de pago:</b> "
                f"{str(cotizacion.forma_pago).replace(chr(10), '<br/>')}",
                normal_left,
            )
        )

    story.append(Spacer(1, 12))
    imagenes = cotizacion.imagenes.all()

    if imagenes:
        story.append(PageBreak())

        story.append(
            Paragraph(
                "5. REGISTRO FOTOGRÁFICO",
                section,
            )
        )

        story.append(Spacer(1, 6))

        for foto in imagenes:
            try:
                ruta_imagen = foto.imagen.path

                img = Image(
                    ruta_imagen,
                    width=95 * mm,
                    height=70 * mm,
                )

                contenido_foto = [
                    [img],
                ]

                if foto.descripcion:
                    contenido_foto.append(
                        [
                            Paragraph(
                                foto.descripcion,
                                normal_left,
                            )
                        ]
                    )

                tabla_foto = Table(
                    contenido_foto,
                    colWidths=[105 * mm],
                )

                tabla_foto.setStyle(
                    TableStyle([
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("BOX", (0, 0), (-1, -1), 0.5, gris_borde),
                        ("TOPPADDING", (0, 0), (-1, -1), 7),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                    ])
                )

                story.append(tabla_foto)
                story.append(Spacer(1, 10))

            except Exception:
                pass

        story.append(Spacer(1, 10))

    # =====================================================
    # CIERRE
    # =====================================================

        elaborado_nombre = (
        cotizacion.elaborado_por.get_full_name().strip()
        or cotizacion.elaborado_por.username
    )

    aprobado_nombre = ""

    if cotizacion.aprobado_por:
        aprobado_nombre = (
            cotizacion.aprobado_por.get_full_name().strip()
            or cotizacion.aprobado_por.username
        )

    firma_realizado = Paragraph(
        f"<b>Realizado por:</b><br/>"
        f"{elaborado_nombre}<br/>"
        f"Supervisor Técnico<br/>"
        f"D&amp;S SOLUCIONES EN BOMBEO S.A.S.",
        normal_left,
    )

    if aprobado_nombre:
        firma_aprobado = Paragraph(
            f"<b>Aprobado por:</b><br/>"
            f"{aprobado_nombre}<br/>"
            f"Coordinador Técnico<br/>"
            f"D&amp;S SOLUCIONES EN BOMBEO S.A.S.",
            normal_left,
        )
    else:
        firma_aprobado = Paragraph(
            "<b>Aprobado por:</b><br/>Pendiente de aprobación",
            normal_left,
        )

    tabla_firmas = Table(
        [[firma_realizado, firma_aprobado]],
        colWidths=[90 * mm, 90 * mm],
    )

    tabla_firmas.setStyle(
        TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
    )

    story.append(Spacer(1, 8))
    story.append(tabla_firmas)
    story.append(Spacer(1, 10))
    story.append(
        Paragraph(
            f"D&amp;S SOLUCIONES EN BOMBEO S.A.S. · {nombre}",
            footer_style,
        )
    )

    doc.build(story)

    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/pdf",
    )
    response["Content-Disposition"] = (
        f'inline; filename="{nombre}.pdf"'
    )

    return response


@login_required
def marcar_cotizacion_enviada(request, cotizacion_id):
    _exigir(
        request.user,
        GRUPO_AUXILIAR,
        GRUPO_COORDINADOR,
        GRUPO_FACTURACION,
        GRUPO_GERENCIA,
        GRUPO_SUPERVISOR,
    )

    cotizacion = get_object_or_404(
        Cotizacion,
        id=cotizacion_id,
    )

    if request.method == "POST" and cotizacion.estado == "LISTA_ENVIAR":
        cotizacion.estado = "ENVIADA"
        cotizacion.fecha_envio = timezone.now()
        cotizacion.save(
            update_fields=[
                "estado",
                "fecha_envio",
                "fecha_actualizacion",
            ]
        )

    return redirect(
        "gestion_comercial:editar_cotizacion",
        cotizacion_id=cotizacion.id,
    )


@login_required
def aprobar_cotizacion(request, cotizacion_id):
    _exigir(
        request.user,
        GRUPO_COORDINADOR,
        GRUPO_GERENCIA,
    )

    cotizacion = get_object_or_404(
        Cotizacion,
        id=cotizacion_id,
    )

    if request.method == "POST" and cotizacion.estado == "ELABORADA":
        ahora = timezone.now()

        cotizacion.estado = "LISTA_ENVIAR"
        cotizacion.fecha_respuesta = ahora
        cotizacion.aprobado_por = request.user
        cotizacion.fecha_aprobacion = ahora

        cotizacion.save(
            update_fields=[
                "estado",
                "fecha_respuesta",
                "aprobado_por",
                "fecha_aprobacion",
                "fecha_actualizacion",
            ]
        )

    return redirect(
        "gestion_comercial:editar_cotizacion",
        cotizacion_id=cotizacion.id,
    )


@login_required
def rechazar_cotizacion(request, cotizacion_id):
    _exigir(
        request.user,
        GRUPO_COORDINADOR,
        GRUPO_GERENCIA,
    )

    cotizacion = get_object_or_404(
        Cotizacion,
        id=cotizacion_id,
    )

    if request.method == "POST" and cotizacion.estado == "ENVIADA":
        cotizacion.estado = "RECHAZADA"
        cotizacion.fecha_respuesta = timezone.now()
        cotizacion.save(
            update_fields=[
                "estado",
                "fecha_respuesta",
                "fecha_actualizacion",
            ]
        )

    return redirect(
        "gestion_comercial:editar_cotizacion",
        cotizacion_id=cotizacion.id,
    )


@login_required
def eliminar_detalle_cotizacion(request, detalle_id):
    _exigir(
        request.user,
        GRUPO_AUXILIAR,
        GRUPO_COORDINADOR,
        GRUPO_FACTURACION,
        GRUPO_GERENCIA,
        GRUPO_SUPERVISOR,
    )

    detalle = get_object_or_404(
        DetalleCotizacion,
        id=detalle_id,
    )
    cotizacion = detalle.cotizacion

    if request.method == "POST":
        detalle.delete()

    return redirect(
        "gestion_comercial:editar_cotizacion",
        cotizacion_id=cotizacion.id,
    )

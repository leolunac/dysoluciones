from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from gestion_comercial.models import CatalogoPrecio


class Command(BaseCommand):

    help = "Importa o actualiza el catálogo de precios desde Excel."

    def add_arguments(self, parser):
        parser.add_argument(
            "archivo",
            type=str,
            help="Ruta del archivo Excel (.xlsx o .xlsm)",
        )

    def handle(self, *args, **options):

        archivo = Path(options["archivo"])

        if not archivo.exists():
            raise CommandError(
                f"No se encontró el archivo: {archivo}"
            )

        self.stdout.write(
            f"Leyendo archivo: {archivo}"
        )

        try:
            libro = load_workbook(
                archivo,
                data_only=True,
                read_only=True,
            )
        except Exception as error:
            raise CommandError(
                f"No fue posible abrir el Excel: {error}"
            )

        if "Precios" not in libro.sheetnames:
            raise CommandError(
                "El archivo no contiene una hoja llamada 'Precios'."
            )

        hoja = libro["Precios"]

        creados = 0
        actualizados = 0
        sin_cambios = 0
        errores = 0

        # Según el archivo actual:
        # A = Código Ítem
        # B = Descripción Ítem
        # C = COSTO UNITARIO

        for numero_fila, fila in enumerate(
            hoja.iter_rows(min_row=2, values_only=True),
            start=2,
        ):

            codigo = fila[0]
            descripcion = fila[1]
            valor = fila[2]

            # Ignorar filas completamente vacías
            if codigo is None and descripcion is None and valor is None:
                continue

            codigo = str(codigo or "").strip()
            descripcion = str(descripcion or "").strip()

            if not codigo:
                self.stderr.write(
                    f"Fila {numero_fila}: código vacío."
                )
                errores += 1
                continue

            if not descripcion:
                self.stderr.write(
                    f"Fila {numero_fila}: descripción vacía "
                    f"para código {codigo}."
                )
                errores += 1
                continue

            try:
                valor_decimal = Decimal(
                    str(valor)
                ).quantize(Decimal("0.01"))

            except (
                InvalidOperation,
                TypeError,
                ValueError,
            ):
                self.stderr.write(
                    f"Fila {numero_fila}: precio inválido "
                    f"para {codigo}: {valor}"
                )
                errores += 1
                continue

            objeto, creado = CatalogoPrecio.objects.get_or_create(
                codigo=codigo,
                defaults={
                    "descripcion": descripcion,
                    "valor": valor_decimal,
                    "activo": True,
                },
            )

            if creado:
                creados += 1
                continue

            cambio = False

            if objeto.descripcion != descripcion:
                objeto.descripcion = descripcion
                cambio = True

            if objeto.valor != valor_decimal:
                objeto.valor = valor_decimal
                cambio = True

            if not objeto.activo:
                objeto.activo = True
                cambio = True

            if cambio:
                objeto.save()
                actualizados += 1
            else:
                sin_cambios += 1

        libro.close()

        self.stdout.write("")
        self.stdout.write(
            self.style.SUCCESS("IMPORTACIÓN TERMINADA")
        )

        self.stdout.write(
            f"Nuevos:       {creados}"
        )

        self.stdout.write(
            f"Actualizados: {actualizados}"
        )

        self.stdout.write(
            f"Sin cambios:  {sin_cambios}"
        )

        self.stdout.write(
            f"Errores:      {errores}"
        )

        self.stdout.write(
            f"Procesados:   "
            f"{creados + actualizados + sin_cambios + errores}"
        )
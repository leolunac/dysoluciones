import re
import unicodedata
from collections import defaultdict
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from operacion.models import Cliente, ContratoPase


# Correspondencias claras entre la base anterior y el nombre actualizado.
# Se usa el nombre anterior normalizado, no el ID, para que el comando sea portable.
COINCIDENCIAS_MANUALES = {
    "42822425": "ARANZAZU ALZATE BEATRIZ",
    "800033658": "CONJUNTO RESIDENCIAL VIZCAYA REAL N.1 P.",
    "800040834": "CONJUNTO RESIDENCIAL BOSQUES DE VIENA",
    "800046300": "CONJUNTO RESIDENCIAL AIRES DE VIGIA DEL FUERTE P. H.",
    "811033278": "REFUGIO DE VILLA VERDE",
    "890938413": "EDIFICIO CENTRO SANTILLANA P.H.",
    "900016568": "UNIDAD RESIDENCIAL SENDERO DE LOS BERNAL",
    "900947614": "EDIFICIO MIRADORES DE ZUÑIGA",
    "901020337": "CONJUNTO DE USO MIXTO CITRINO",
    "901248938": "CONJUNTO RESIDENCIAL URBANIZACION ROSSET",
    "901469538": "PASEO COMERCIAL AMSTERDAM PLAZA - P.H",
    "901649721": "6 GROUP",
    "901661339": "URBANIZACION SAINT MICHEL APARTAMENTOS PROPIEDAD HORIZONTAL",
    "901776971": "URBANIZACION NATAL",
}


def limpiar_nit(valor):
    if valor is None:
        return ""
    texto = str(valor).strip()
    if texto.endswith(".0"):
        texto = texto[:-2]
    return re.sub(r"\D", "", texto)


def normalizar(texto):
    texto = unicodedata.normalize("NFKD", str(texto or ""))
    texto = texto.encode("ascii", "ignore").decode("ascii").upper()
    return re.sub(r"[^A-Z0-9]", "", texto)


def texto_limpio(valor):
    return " ".join(str(valor or "").strip().split())


def decimal_cuota(valor, fila):
    try:
        return Decimal(str(valor)).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise CommandError(
            f"Valor de cuota invalido en Contratos, fila {fila}: {valor!r}"
        ) from exc


class Command(BaseCommand):
    help = (
        "Importa la base actualizada de clientes y contratos/pases desde Excel. "
        "Sin --aplicar solo realiza una simulacion."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "archivo",
            nargs="?",
            default="Relacion Unidades - Clientes.xlsx",
            help="Ruta del archivo Excel",
        )
        parser.add_argument(
            "--aplicar",
            action="store_true",
            help="Guarda los cambios en la base de datos",
        )

    def handle(self, *args, **options):
        ruta = Path(options["archivo"])
        aplicar = options["aplicar"]

        if not ruta.exists():
            raise CommandError(f"No existe el archivo: {ruta}")

        try:
            libro = load_workbook(ruta, data_only=True, read_only=True)
        except Exception as exc:
            raise CommandError(f"No fue posible abrir el Excel: {exc}") from exc

        requeridas = {"clientes", "Contratos"}
        faltantes = requeridas.difference(libro.sheetnames)
        if faltantes:
            raise CommandError(
                "Faltan hojas requeridas: " + ", ".join(sorted(faltantes))
            )

        clientes_excel = self.leer_clientes(libro["clientes"])
        pases_excel = self.leer_pases(libro["Contratos"])
        libro.close()

        self.validar_archivo(clientes_excel, pases_excel)
        plan = self.construir_plan(clientes_excel)

        self.mostrar_simulacion(clientes_excel, pases_excel, plan)

        if plan["ambiguos"]:
            raise CommandError(
                "Hay coincidencias ambiguas. No se puede aplicar hasta resolverlas."
            )

        if not aplicar:
            self.stdout.write(
                self.style.WARNING(
                    "SIMULACION: no se modifico la base de datos. "
                    "Use --aplicar solo despues de revisar este informe."
                )
            )
            return

        with transaction.atomic():
            resultado = self.aplicar_importacion(clientes_excel, pases_excel, plan)

        self.stdout.write(self.style.SUCCESS("IMPORTACION COMPLETADA"))
        for etiqueta, valor in resultado.items():
            self.stdout.write(f"{etiqueta}: {valor}")

    def leer_clientes(self, hoja):
        registros = []
        for numero_fila, fila in enumerate(hoja.iter_rows(values_only=True), start=1):
            nit = limpiar_nit(fila[0] if len(fila) > 0 else None)
            nombre = texto_limpio(fila[1] if len(fila) > 1 else None)
            if not nit and not nombre:
                continue
            if not nit or not nombre:
                raise CommandError(
                    f"Cliente incompleto en hoja clientes, fila {numero_fila}"
                )
            registros.append(
                {"fila": numero_fila, "nit": nit, "nombre": nombre}
            )
        return registros

    def leer_pases(self, hoja):
        mapa_periodicidad = {
            "MENSUAL": "MENSUAL",
            "BIMESTRAL": "BIMESTRAL",
            "TRIMESTRAL": "TRIMESTRAL",
            "CADA4MESES": "CUATRIMESTRAL",
            "CUATRIMESTRAL": "CUATRIMESTRAL",
        }
        registros = []
        for numero_fila, fila in enumerate(
            hoja.iter_rows(min_row=3, values_only=True), start=3
        ):
            nit = limpiar_nit(fila[1] if len(fila) > 1 else None)
            nombre_servicio = texto_limpio(fila[2] if len(fila) > 2 else None)
            numero_pase = texto_limpio(fila[3] if len(fila) > 3 else None).upper()
            valor = fila[4] if len(fila) > 4 else None
            frecuencia_original = texto_limpio(
                fila[5] if len(fila) > 5 else None
            ).upper()
            direccion = texto_limpio(fila[6] if len(fila) > 6 else None)

            if not any((nit, nombre_servicio, numero_pase, valor, frecuencia_original)):
                continue
            if not all((nit, nombre_servicio, numero_pase, frecuencia_original)):
                raise CommandError(
                    f"Contrato incompleto en hoja Contratos, fila {numero_fila}"
                )

            clave_frecuencia = normalizar(frecuencia_original)
            periodicidad = mapa_periodicidad.get(clave_frecuencia)
            if not periodicidad:
                raise CommandError(
                    f"Periodicidad desconocida en fila {numero_fila}: "
                    f"{frecuencia_original}"
                )

            registros.append(
                {
                    "fila": numero_fila,
                    "nit": nit,
                    "nombre_servicio": nombre_servicio,
                    "numero_pase": numero_pase,
                    "valor_cuota": decimal_cuota(valor, numero_fila),
                    "periodicidad": periodicidad,
                    "direccion_servicio": direccion,
                }
            )
        return registros

    def validar_archivo(self, clientes, pases):
        nits = [x["nit"] for x in clientes]
        nits_repetidos = sorted({x for x in nits if nits.count(x) > 1})
        if nits_repetidos:
            raise CommandError(
                "NIT repetidos en clientes: " + ", ".join(nits_repetidos)
            )

        numeros = [x["numero_pase"] for x in pases]
        pases_repetidos = sorted({x for x in numeros if numeros.count(x) > 1})
        if pases_repetidos:
            raise CommandError(
                "Pases repetidos: " + ", ".join(pases_repetidos)
            )

        nits_clientes = set(nits)
        contratos_sin_cliente = sorted(
            {x["nit"] for x in pases if x["nit"] not in nits_clientes}
        )
        if contratos_sin_cliente:
            raise CommandError(
                "Hay contratos cuyo NIT no existe en clientes: "
                + ", ".join(contratos_sin_cliente)
            )

    def construir_plan(self, clientes_excel):
        clientes_db = list(Cliente.objects.all().order_by("id"))
        por_nit = {x.nit: x for x in clientes_db if x.nit}
        por_nombre = defaultdict(list)
        for cliente in clientes_db:
            por_nombre[normalizar(cliente.nombre)].append(cliente)

        plan = {
            "por_nit": [],
            "por_nombre": [],
            "manuales": [],
            "nuevos": [],
            "ambiguos": [],
        }

        for registro in clientes_excel:
            nit = registro["nit"]
            nombre = registro["nombre"]

            if nit in por_nit:
                plan["por_nit"].append((registro, por_nit[nit]))
                continue

            exactos = por_nombre.get(normalizar(nombre), [])
            if len(exactos) == 1:
                plan["por_nombre"].append((registro, exactos[0]))
                continue
            if len(exactos) > 1:
                plan["ambiguos"].append((registro, exactos))
                continue

            nombre_anterior = COINCIDENCIAS_MANUALES.get(nit)
            if nombre_anterior:
                candidatos = por_nombre.get(normalizar(nombre_anterior), [])
                if len(candidatos) == 1:
                    plan["manuales"].append((registro, candidatos[0]))
                    continue
                if len(candidatos) > 1:
                    plan["ambiguos"].append((registro, candidatos))
                    continue

            plan["nuevos"].append(registro)

        return plan

    def mostrar_simulacion(self, clientes, pases, plan):
        nits_con_pase = {x["nit"] for x in pases}
        frecuencias = defaultdict(int)
        for pase in pases:
            frecuencias[pase["periodicidad"]] += 1

        self.stdout.write("=" * 64)
        self.stdout.write("SIMULACION DE IMPORTACION")
        self.stdout.write("=" * 64)
        self.stdout.write(f"Clientes en Excel: {len(clientes)}")
        self.stdout.write(f"Coincidencias por NIT: {len(plan['por_nit'])}")
        self.stdout.write(
            f"Coincidencias exactas por nombre: {len(plan['por_nombre'])}"
        )
        self.stdout.write(
            f"Coincidencias manuales verificables: {len(plan['manuales'])}"
        )
        self.stdout.write(f"Clientes que se crearian: {len(plan['nuevos'])}")
        self.stdout.write(f"Coincidencias ambiguas: {len(plan['ambiguos'])}")
        self.stdout.write(f"Pases en Excel: {len(pases)}")
        self.stdout.write(f"Clientes con pase: {len(nits_con_pase)}")
        self.stdout.write(
            f"Clientes del Excel sin pase: {len(clientes) - len(nits_con_pase)}"
        )
        self.stdout.write(
            "Periodicidades: "
            + ", ".join(
                f"{nombre}={cantidad}"
                for nombre, cantidad in sorted(frecuencias.items())
            )
        )

        if plan["manuales"]:
            self.stdout.write("\nCOINCIDENCIAS MANUALES PROPUESTAS:")
            for registro, cliente in plan["manuales"]:
                self.stdout.write(
                    f"  {registro['nit']} | {registro['nombre']} "
                    f"=> ID {cliente.id} | {cliente.nombre}"
                )

        if plan["nuevos"]:
            self.stdout.write("\nCLIENTES QUE SE CREARIAN:")
            for registro in plan["nuevos"]:
                self.stdout.write(
                    f"  {registro['nit']} | {registro['nombre']}"
                )

        if plan["ambiguos"]:
            self.stdout.write("\nCOINCIDENCIAS AMBIGUAS:")
            for registro, candidatos in plan["ambiguos"]:
                lista = "; ".join(
                    f"ID {x.id}: {x.nombre}" for x in candidatos
                )
                self.stdout.write(
                    f"  {registro['nit']} | {registro['nombre']} => {lista}"
                )

        pase_provisional = next(
            (x for x in pases if x["numero_pase"] == "PASE259"), None
        )
        if pase_provisional:
            self.stdout.write(
                "\nPASE259 se importara con valor "
                f"{pase_provisional['valor_cuota']} y estado PROVISIONAL."
            )

    def aplicar_importacion(self, clientes_excel, pases_excel, plan):
        clientes_por_nit = {}
        actualizados = 0
        creados = 0

        relaciones = []
        relaciones.extend(plan["por_nit"])
        relaciones.extend(plan["por_nombre"])
        relaciones.extend(plan["manuales"])

        for registro, cliente in relaciones:
            cliente.nit = registro["nit"]
            cliente.nombre = registro["nombre"]
            cliente.activo = True
            cliente.save(update_fields=["nit", "nombre", "activo"])
            clientes_por_nit[registro["nit"]] = cliente
            actualizados += 1

        for registro in plan["nuevos"]:
            cliente = Cliente.objects.create(
                nit=registro["nit"],
                nombre=registro["nombre"],
                direccion="",
                telefono_porteria="",
                administrador="",
                email="",
                tipo_contrato="SIN_CONTRATO",
                frecuencia_lavado=6,
                activo=True,
            )
            clientes_por_nit[registro["nit"]] = cliente
            creados += 1

        nits_con_pase = {x["nit"] for x in pases_excel}
        ids_con_pase = {
            clientes_por_nit[nit].id for nit in nits_con_pase
        }

        # Los clientes nunca se desactivan. Los que no aparecen con pase quedan
        # como SIN_CONTRATO. Si un cliente con pase ya era 7X24, se conserva.
        Cliente.objects.exclude(id__in=ids_con_pase).update(
            tipo_contrato="SIN_CONTRATO"
        )
        Cliente.objects.filter(id__in=ids_con_pase).exclude(
            tipo_contrato="7X24"
        ).update(tipo_contrato="PREVENTIVO")

        pases_creados = 0
        pases_actualizados = 0
        numeros_vigentes = []

        for registro in pases_excel:
            numero = registro["numero_pase"]
            numeros_vigentes.append(numero)
            provisional = numero == "PASE259"
            valores = {
                "cliente": clientes_por_nit[registro["nit"]],
                "nombre_servicio": registro["nombre_servicio"],
                "tipo": "PREVENTIVO",
                "valor_cuota": registro["valor_cuota"],
                "periodicidad": registro["periodicidad"],
                "estado_valor": "PROVISIONAL" if provisional else "DEFINITIVO",
                "direccion_servicio": registro["direccion_servicio"],
                "observaciones": (
                    "Pendiente de aprobacion del presupuesto del cliente."
                    if provisional
                    else ""
                ),
                "dia_limite_facturacion": 15,
                "activo": True,
            }
            _, creado = ContratoPase.objects.update_or_create(
                numero_pase=numero,
                defaults=valores,
            )
            if creado:
                pases_creados += 1
            else:
                pases_actualizados += 1

        pases_desactivados = ContratoPase.objects.exclude(
            numero_pase__in=numeros_vigentes
        ).update(activo=False)

        return {
            "Clientes actualizados": actualizados,
            "Clientes creados": creados,
            "Pases creados": pases_creados,
            "Pases actualizados": pases_actualizados,
            "Pases anteriores desactivados": pases_desactivados,
            "Total clientes en la base": Cliente.objects.count(),
            "Total pases activos": ContratoPase.objects.filter(activo=True).count(),
            "Fecha": date.today().isoformat(),
        }

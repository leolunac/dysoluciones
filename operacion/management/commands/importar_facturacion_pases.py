import calendar
from collections import Counter
from datetime import date
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from operacion.models import ContratoPase, FacturacionPase


MESES = {
    10: 1, 11: 2, 12: 3, 13: 4, 14: 5, 15: 6,
    16: 7, 17: 8, 18: 9, 19: 10, 20: 11, 21: 12,
}

VALORES_VACIOS = {"", "0", "NO", "N", "FALSE", "FALSO", "-", "--"}


def texto(valor):
    return str(valor or "").strip()


def esta_marcado(valor):
    if valor is None:
        return False
    if isinstance(valor, bool):
        return valor
    return texto(valor).upper() not in VALORES_VACIOS


class Command(BaseCommand):
    help = "Importa la programacion mensual de facturacion de los pases desde Excel."

    def add_arguments(self, parser):
        parser.add_argument(
            "archivo",
            nargs="?",
            default="Relacion Unidades - Clientes.xlsx",
            help="Ruta del archivo Excel.",
        )
        parser.add_argument("--anio", type=int, default=2026)
        parser.add_argument(
            "--aplicar",
            action="store_true",
            help="Guarda los registros. Sin esta opcion solo simula.",
        )

    def handle(self, *args, **options):
        ruta = Path(options["archivo"])
        anio = options["anio"]
        aplicar = options["aplicar"]

        if not ruta.exists():
            raise CommandError(f"No existe el archivo: {ruta}")

        libro = load_workbook(ruta, data_only=True, read_only=True)
        if "Contratos" not in libro.sheetnames:
            raise CommandError("No existe la hoja 'Contratos'.")
        hoja = libro["Contratos"]

        contratos = {
            c.numero_pase.strip().upper(): c
            for c in ContratoPase.objects.select_related("cliente").all()
        }

        hoy = timezone.localdate()
        candidatos = []
        pases_no_encontrados = set()
        filas_sin_pase = []
        duplicados_excel = []
        vistos = set()
        por_mes = Counter()
        por_estado = Counter()

        for fila in range(3, hoja.max_row + 1):
            numero = texto(hoja.cell(fila, 4).value).upper()
            if not numero:
                if any(esta_marcado(hoja.cell(fila, col).value) for col in MESES):
                    filas_sin_pase.append(fila)
                continue

            contrato = contratos.get(numero)
            if contrato is None:
                if any(esta_marcado(hoja.cell(fila, col).value) for col in MESES):
                    pases_no_encontrados.add(numero)
                continue

            for columna, mes in MESES.items():
                if not esta_marcado(hoja.cell(fila, columna).value):
                    continue

                periodo = date(anio, mes, 1)
                clave = (contrato.id, periodo)
                if clave in vistos:
                    duplicados_excel.append(f"{numero} / {periodo:%Y-%m}")
                    continue
                vistos.add(clave)

                dia = min(
                    contrato.dia_limite_facturacion or 15,
                    calendar.monthrange(anio, mes)[1],
                )
                fecha_limite = date(anio, mes, dia)
                estado = (
                    "POR_CONFIRMAR"
                    if periodo < hoy.replace(day=1)
                    else "PENDIENTE"
                )
                observacion = ""
                if contrato.estado_valor == "PROVISIONAL":
                    observacion = (
                        "Valor provisional: validar el presupuesto aprobado "
                        "antes de registrar la factura de Siigo."
                    )

                candidatos.append({
                    "contrato": contrato,
                    "periodo": periodo,
                    "fecha_limite": fecha_limite,
                    "valor_programado": Decimal(contrato.valor_cuota),
                    "estado": estado,
                    "observaciones": observacion,
                })
                por_mes[periodo.strftime("%Y-%m")] += 1
                por_estado[estado] += 1

        libro.close()

        existentes = {
            (x.contrato_id, x.periodo): x
            for x in FacturacionPase.objects.filter(periodo__year=anio)
        }
        nuevos = []
        actualizables = []
        protegidos = []

        for datos in candidatos:
            clave = (datos["contrato"].id, datos["periodo"])
            actual = existentes.get(clave)
            if actual is None:
                nuevos.append(datos)
            elif actual.estado == "FACTURADA":
                protegidos.append(actual)
            else:
                actualizables.append((actual, datos))

        self.stdout.write("=" * 64)
        self.stdout.write("SIMULACION DE PROGRAMACION DE FACTURACION")
        self.stdout.write("=" * 64)
        self.stdout.write(f"Archivo: {ruta}")
        self.stdout.write(f"Año: {anio}")
        self.stdout.write(f"Marcas mensuales encontradas: {len(candidatos)}")
        self.stdout.write(f"Registros nuevos: {len(nuevos)}")
        self.stdout.write(f"Registros existentes actualizables: {len(actualizables)}")
        self.stdout.write(f"Facturas existentes protegidas: {len(protegidos)}")
        self.stdout.write(
            "Estados: " + ", ".join(f"{k}={v}" for k, v in sorted(por_estado.items()))
        )
        self.stdout.write(
            "Meses: " + ", ".join(f"{k}={v}" for k, v in sorted(por_mes.items()))
        )

        if pases_no_encontrados:
            self.stdout.write(self.style.ERROR(
                "Pases no encontrados: " + ", ".join(sorted(pases_no_encontrados))
            ))
        if filas_sin_pase:
            self.stdout.write(self.style.ERROR(
                "Filas marcadas sin numero de pase: " + ", ".join(map(str, filas_sin_pase))
            ))
        if duplicados_excel:
            self.stdout.write(self.style.ERROR(
                "Duplicados en Excel: " + ", ".join(duplicados_excel)
            ))

        if pases_no_encontrados or filas_sin_pase or duplicados_excel:
            raise CommandError("La validacion detecto inconsistencias; no se guardo nada.")

        if not aplicar:
            self.stdout.write(self.style.WARNING(
                "SIMULACION: no se modifico la base. Use --aplicar despues de revisar."
            ))
            return

        with transaction.atomic():
            FacturacionPase.objects.bulk_create([
                FacturacionPase(**datos) for datos in nuevos
            ])

            for actual, datos in actualizables:
                actual.fecha_limite = datos["fecha_limite"]
                actual.valor_programado = datos["valor_programado"]
                actual.estado = datos["estado"]
                if datos["observaciones"] and not actual.observaciones:
                    actual.observaciones = datos["observaciones"]
                actual.save(update_fields=[
                    "fecha_limite",
                    "valor_programado",
                    "estado",
                    "observaciones",
                    "actualizado",
                ])

        self.stdout.write(self.style.SUCCESS("IMPORTACION COMPLETADA"))
        self.stdout.write(f"Creados: {len(nuevos)}")
        self.stdout.write(f"Actualizados: {len(actualizables)}")
        self.stdout.write(f"Facturadas preservadas: {len(protegidos)}")
        self.stdout.write(
            f"Total programado para {anio}: "
            f"{FacturacionPase.objects.filter(periodo__year=anio).count()}"
        )
